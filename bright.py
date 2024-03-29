# github.com/freddiexyz/bright

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from datetime import date, datetime

import _secret
import config

import logging
import records
import pysftp
import re


class Database(records.Database):

    def __init__(self):
        super().__init__('mysql+pymysql://{userpass}@{host}/{database}'.format(**_secret.db_login))


class Claim():

    def __init__(self, patient, procedures, carrier):
        self.patient = patient
        self.procedures = procedures
        self.carrier = carrier

        self.patient['patient name'] = self.patient['first_name'] + ' ' + self.patient['last_name']
        self.patient['birthdate'] = self.patient['birthdate'].strftime("%d%m%Y")
        self.patient['prov_city'] = 'Christchurch'

        self.is_pa = bool(self.patient['claimform'] == self.carrier['pa_claimform'])
        if not self.is_pa:
            del self.patient['prior_approval']

        if self.carrier is config.OHSA:
            self.patient['decile'] = get_decile(self.patient['school'])
        elif self.carrier is config.SDSC:
            if self.patient['subnum'] == '.':
                self.tx_tickbox_num = 3
            else:
                self.tx_tickbox_num = 1

        for proc in self.procedures:
            proc['date'] = proc['proc_date'].strftime("%d.%m.%y")
            if proc['teeth']:
                proc['teeth'] = ','.join(config.teeth[tooth] for tooth in proc['teeth'].split(',')\
                 if tooth.isalnum()) # conversion to not american notation

        self.patient['fee'] = self.calculate_fee() # sum of procedures
        self.summary = None

    def __len__(self):
        return len(self.procedures)

    def __str__(self):
        return f"{self.patient['last_name']}, {self.patient['first_name']}\n\t" +\
            "\n\t".join("{proc_date} | {code} | {quantity} | {fee} | {teeth}".format(
                **proc) for proc in self.procedures)

    @classmethod
    def merge(cls, patients, procedures, carrier):
        proc_index = 0
        for patient in patients:
            patient_procs = [procedures[proc_index]]
            proc_index += 1
            try:
                while procedures[proc_index]['claimnum'] == patient['claimnum']:
                    patient_procs.append(procedures[proc_index])
                    proc_index += 1
            except IndexError: # exhausted all procedures
                pass
            yield cls(patient, patient_procs, carrier)

    @classmethod
    def from_waiting(cls, db, carrier):
        patients = db.query(config.SELECT_PATIENTS_WAITING.format(**carrier)).all(as_dict=True)
        procedures = db.query(config.SELECT_PROCEDURES_WAITING.format(**carrier)).all(as_dict=True)

        return cls.merge(patients, procedures, carrier)

    @classmethod
    def from_waiting_pa(cls, db, carrier):
        plannum = carrier['plannum']
        claimform = carrier['pa_claimform']
        patients = db.query(config.SELECT_PATIENTS_WAITING.format(plannum=planum, claimform=claimform)).all(as_dict=True)
        procedures = db.query(config.SELECT_PROCEDURES_WAITING.format(plannum=planum, claimform=claimform)).all(as_dict=True)

        return cls.merge(patients, procedures, carrier)

    @classmethod
    def from_claimnum(cls, db, carrier, name):
        patients = db.query(config.SELECT_PATIENTS_SENT.format(name)).all(as_dict=True)
        procedures = db.query(config.SELECT_PROCEDURES_SENT.format(name)).all(as_dict=True)

        return cls.merge(patients, procedures, carrier)

    def calculate_fee(self):
        fee = 0
        for proc in self.procedures:
            fee += float(proc['fee'])
            proc['fee'] = f"{proc['fee']:6.2f}" # 2 decimal place float since currency
        return fee

    def validate(self):
        missing_info = []
        if not self.validate_nhi():
            missing_info.append('NHI')
        if self.carrier['name'] == 'SDSC' and not self.validate_cds_ref():
            missing_info.append('subnum')
        elif self.carrier['name'] == 'OHSA':
            pass #TODO: more ohsa stuff
        if self.is_pa and not self.validate_pa():
            missing_info.append('prior_approval')
        missing_info.extend(field for field in self.patient if not self.patient[field])
        if missing_info:
            self.missing_info = missing_info
        return not missing_info

    def validate_nhi(self):
        if not check_nhi(self.patient['NHI']):
            return False
        return True

    def validate_cds_ref(self):
        if not check_cds_ref(self.patient['subnum']):
            return False
        return True
        
    def validate_pa(self):
        # could pattern match PA numbers in future
        if not self.patient['prior_approval']:
            return False
        return True

    def to_form(self, cvs):
        form_length = self.carrier['form_length_pa'] if self.is_pa else self.carrier['form_length']
        for page_num in range(((len(self) - 1) // form_length) + 1):
            procs = self.procedures[page_num*form_length:(page_num+1)*form_length] # what
            self.to_page(cvs, procs)

    def to_page(self, cvs, procs):
        width, height = A4
        form_coords = self.carrier['form_coords']

        # Form image and patient info
        cvs.setFont('Courier', 12) # courier since monospaced
        cvs.drawImage(self.carrier['form_img'], 0,0, width=width, height=height) # fullpage image

        pat_coords = form_coords['patient']
        if self.is_pa:
            proc_coords = form_coords['procedures_pa']
            draw(cvs, self.patient['prior_approval'],
                *form_coords['prior_approval'])
        else:
            proc_coords = form_coords['procedures']

        for field, coords in pat_coords.items():
            draw(cvs, self.patient[field], *coords)

        #tickboxes
        cvs.setFont('Helvetica-Bold', 23)

        if self.patient['gender'] == 1: # female
            gender_tickbox_coords = form_coords['tickboxes']['gender_f']
        elif self.patient['gender'] == 2: # male
            gender_tickbox_coords = form_coords['tickboxes']['gender_m']

        draw(cvs, 'x', *gender_tickbox_coords)

        tx_tickbox_coords = form_coords['tickboxes'][f'tx{self.tx_tickbox_num}']
        draw(cvs, 'x', *tx_tickbox_coords)

        # procedures + fee total
        cvs.setFont('Courier', 10) # so it will fit
        for num, proc in enumerate(procs):
            for field, coords in proc_coords.items():
                # lowers height of each line by set amount each procedure
                draw(cvs, str(proc[field]), coords[0], coords[1]-(num*config.proc_line_step))

        draw(cvs, self.patient['fee'], *form_coords['total'])
        cvs.showPage()

    def update_claimstatus(self, status):
        assert status in ('S', 'R', 'W', 'H') # sent, recieved, waiting, hold
        print(config.UPDATE_CLAIMSTATUS.format(status, self.patient['claimnum']))
        # self.summary.db.query(config.UPDATE_CLAIMSTATUS.format(status, claimnums))

    def remove_from_sentclaim(self):
        print(config.DELETE_SENTCLAIM.format(self.patient['claimnum']))
        # self.summary.db.query(config.DELETE_SENTCLAIM.format(self.patient['claimnum']))

    def remove_procedure(self, procedure):
        assert procedure in self.procedures
        if len(self) == 1:
            self.summary.remove_claim(self)
            return
        self.summary.db.query(config.DELETE_CLAIMPROC.format(procnum=procedure['procnum'],
            claimnum=self.patient['claimnum']))
        self.procedures.remove(procedure)

        self.calculate_fee()

        
class Summary():
    def __init__(self, claims, carrier, name=None):
        self.claims = list(claims)
        self.carrier = carrier

        self.calculate_totals()

        for claim in self.claims:
            claim.summary = self

        self.claimnums = ','.join(str(claim.patient['claimnum']) for claim in self.claims)

        if name is None:
            self.name = date.today().strftime(f'{self.carrier["name"]}%d%m%y')
        else:
            self.name = name

    def __len__(self):
        return len(self.claims)

    def __str__(self):
        return f'Number of Patients: {len(self)}' +\
            f'\nTotal: {self.total:>8.2f}' +\
            f'\nGST: {self.GST:>8.2f}' +\
            f'\nTotal inc GST: {self.total_inc_GST:>8.2f}'

    @classmethod
    def from_waiting(cls, db, carrier, name=None):
        cls.db = db
        claims = []
        gen = Claim.from_waiting(db, carrier)
        try:
            while len(claims) < config.MAX_CLAIMS:
                claim = next(gen)
                if claim.validate():
                    claims.append(claim)
                else:
                    pass # TODO: deal with claims needing info
        except StopIteration:
            pass
        return cls(claims, carrier, name)

    @classmethod
    def from_sentclaim(cls, db, carrier, name):
        cls.db = db
        claims = list(Claim.from_claimnum(db, carrier, name))
        return cls(claims, carrier, name)

    def calculate_totals(self):
        self.total = sum(claim.patient['fee'] for claim in self.claims)
        self.GST = self.total * config.GST
        self.total_inc_GST = self.total + self.GST

    def to_forms(self, cvs, summary=True):
        if summary:
            self.to_summary(cvs)
        for claim in self.claims:
            claim.to_form(cvs)
        return cvs

    def to_summary(self, cvs):
        width, height = A4
        cvs.drawImage(self.carrier['summary_img'], 0,0, width=width, height=height)
        cvs.setFont('Courier', 12)

        for field, value in _secret.practice_details.items():
            draw(cvs, value, *config.summary_coords[field])

        draw(cvs, self.name,          *config.summary_coords['claim_reference'])
        draw(cvs, len(self),          *config.summary_coords['num_patients'])
        draw(cvs, self.total_inc_GST, *config.summary_coords['fee_inc'])
        draw(cvs, self.total,         *config.summary_coords['fee_ex'])
        draw(cvs, self.GST,           *config.summary_coords['GST'])

        cvs.showPage()
        return cvs

    def to_spreadsheet(self, filename):
        wb = Workbook()
        ws = wb.active
        ws['C1'] = filename
        
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        
        columns = {
            'A' : 'claimnum',
            'B' : 'NHI',
            'C' : 'patient name',
            'D' : 'claimdate',
            'E' : 'fee',
        }

        for column, field in columns.items():
            ws[f'{column}2'] = field.title()

        for num, claim in enumerate(self.claims):
            for column, field in columns.items():
                ws[f'{column}{num + 3}'] = claim.patient[field]
                if column == 'E':
                    ws[f'{column}{num + 3}'].style = 'Currency'

        table = Table(displayName="Table1", ref=f'A2:E{len(self) + 2}')
        style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        for num, value in enumerate((self.total, self.GST, self.total_inc_GST)):
            ws[f'E{len(self) + num + 3}'] = value
            ws[f'E{len(self) + num + 3}'].style = 'Currency'

        wb.save(f'test_output\\{filename}.xlsx')

    def update_claimstatus(self, status):
        assert status in ('S', 'R', 'W', 'H') # sent, recieved, waiting, hold
        print(config.UPDATE_CLAIMSTATUS.format(status, self.claimnums))

    def update_datesent(self):
        print(config.UPDATE_DATESENT.format(date.today(), self.claimnums))

    def insert_to_sentclaim(self):
        # ??????
        claimnums = ',\n'.join(f"({claim.patient['claimnum']}, {self.name}, {self.carrier['name']})" for claim in self.claims)
        query_string = config.INSERT_SENTCLAIM + claimnums
        print(query_string)

    def remove_from_sentclaim(self):
        print(config.DELETE_SENTCLAIM.format(self.claimnums))

    def send(self, forms=True, summary=True, spreadsheet=True):
        self.update_claimstatus('S')
        self.insert_to_sentclaim()
        self.update_datesent()

        self.insert_task()
        self.insert_tasknote(f'Sent\n{self}')

        if forms:
            self.to_forms(f'{self.name}_forms')
        if summary:
            self.to_summary(f'{self.name}_summary')
        if spreadsheet:
            self.to_spreadsheet(f'{self.name}_spreadsheet')

    def receive(self):
        self.update_claimstatus('R')
        self.insert_tasknote(f'Received\n{self}')
        self.update_task()
        print(config.UPDATE_SENTCLAIM.format(self.claimnums))
        print(config.INSERT_CLAIMPAYMENT.format( # needs testing
            carrier = self.carrier['name'],
            date    = date.today(),
            amount  = self.total,
            claims  = self.claimnums,
            note    = self.name))

    def remove_claim(self, claim):
        claim.update_claimstatus('W')
        claim.remove_from_sentclaim()
        claim.summary = None

        self.claims.remove(claim)
        self.claimnums = ','.join(str(claim.patient['claimnum']) for claim in self.claims)
        self.calculate_totals()

    def remove_procedure(self, claim, procedure):
        claim.remove_procedure(procedure)
        self.calculate_totals()

    def insert_task(self):
        assert self.get_tasknum() is None
        self.db.query(config.INSERT_TASK.format(descript=self.name, datetime=datetime.now()))

    def get_tasknum(self):
        res = self.db.query(config.SELECT_TASKNUM.format(descript=self.name)).one(as_dict=True)
        if res:
            return res['TaskNum']
        else:
            return None

    def insert_tasknote(self, note):
        self.db.query(config.INSERT_TASKNOTE.format(tasknum=self.get_tasknum(), datetime=datetime.now(), note=note))

    def update_task(self):
        self.db.query(config.UPDATE_TASK.format(tasknum=self.get_tasknum()))

    def remove_task(self):
        self.db.query(config.DELETE_TASK.format(tasknum=self.get_tasknum()))


def draw(cvs, value, *coords):
    # wrapper for reportlabs.canvas.Canvas.drawString
    if type(value) is float:
        value = f'{value:0.2f}'
    if coords[2:]:
        cvs.drawString(coords[0], coords[1], str(value), **coords[2])
    else:
        cvs.drawString(coords[0], coords[1], str(value))

def check_nhi(nhi):
    '''Uses the check digit to verify that NHI is valid'''
    alpha = 'ABCDEFGHJKLMNPQRSTUVWXYZ' # does not contain 'I' or 'O'
    nhi = str(nhi).upper().strip()
    if not re.fullmatch(f'^[{alpha}]{{3}}[\d]{{4}}$', nhi): # matches pattern AAANNNN
        return False
    # Parity check maths
    sum_letters = sum((str.find(alpha, str(nhi[i])) + 1) * (7-i) for i in range(3))
    sum_numbers = sum(int(nhi[i]) * (7-i) for i in range(3,6))
    total = sum_letters + sum_numbers
    check = 11 - (total % 11)
    if check == 11 or check % 10 != int(nhi[6]):
        return False
    return True

def check_cds_ref(ref_num):
    '''Matches referral number to known patterns
    ######-[SED/SBD] or SED-049-#### or SED17[NHI]'''
    ref_num = ref_num.strip().upper()
    if ref_num == '.':
        return True
    if re.fullmatch('^\d{6}[-|\s]?[A-Z]{3}.*', ref_num):
        return True
    if re.fullmatch('^SED-\d{3}-\d{4}$', ref_num):
        return True
    if re.fullmatch('^(SED|SDB)\d{2}[-|\s]?[A-Z]{3}\d{4}$', ref_num):
        return True
    return False

def get_decile(pat_school):
    '''Attempts to match a given school to known schools'''
    for school in config.schools:
        if any(re.fullmatch(f"^{pattern}.*$", pat_school.upper().strip()) for pattern in school):
            return config.schools[school]
    return 0 # placeholder for not found or N/A


if __name__ == '__main__':
    name = 'BSTEST'
    with Database() as db:
        test_summary = Summary.from_waiting(db, config.SDSC, name)

    filename = config.FILENAME_FORMAT.format(
                    payee_num    = _secret.practice_details['payee_number'],
                    claim_ref    = name,
                    claim_type   = test_summary.carrier['name'],
                    num_patients = len(test_summary)
                    )

    cvs = canvas.Canvas(f'.\\test_output\\{filename}.pdf', pagesize=A4)
    test_summary.to_forms(cvs, summary=False)
    cvs.save()